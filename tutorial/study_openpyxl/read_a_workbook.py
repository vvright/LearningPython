from openpyxl import load_workbook

wb = load_workbook(filename='empty_book.xlsx')
sheet_ranges = wb['range names']
# print(sheet_ranges['D18'].value)

for row in sheet_ranges.rows:
    for cell in row:
        print(cell, cell.value)

print(sheet_ranges.max_row, sheet_ranges.max_column)